"""Gera planilha de custos e roteiro de 5 dias para viagem a Paris."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

OUTPUT = "Custo_Viagem_Paris.xlsx"
CAMBIO_CELL = "G3"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
DAY_FILLS = [
    PatternFill("solid", fgColor="D6E4F0"),
    PatternFill("solid", fgColor="E2EFDA"),
    PatternFill("solid", fgColor="FCE4D6"),
    PatternFill("solid", fgColor="E4DFEC"),
    PatternFill("solid", fgColor="FFF2CC"),
]
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
TOTAL_FILL = PatternFill("solid", fgColor="E2EFDA")
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="B4B4B4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
EUR_FMT = '#,##0.00 "€"'
BRL_FMT = "R$ #,##0.00"
LINK_FONT = Font(color="0563C1", underline="single")

# Links oficiais de reserva e informações
BOOKING_LINKS = [
    {
        "categoria": "Transporte",
        "atracao": "Navigo semanal (zonas 1–5)",
        "url": "https://www.iledefrance-mobilites.fr/en/tickets-fares/detail/navigo-weekly-ticket",
        "como_agendar": "Compre em estação (a partir de sexta p/ semana seguinte) ou app Île-de-France Mobilités",
        "antecedencia": "Válido seg–dom; carregue no cartão Navigo Découverte ou celular",
    },
    {
        "categoria": "Transporte",
        "atracao": "Cartão Navigo Découverte",
        "url": "https://www.iledefrance-mobilites.fr/en/tickets-fares/media/navigo-decouverte",
        "como_agendar": "Compre em guichês de metrô/RER; leve foto 3×4 e documento",
        "antecedencia": "No dia — necessário antes de carregar o passe semanal",
    },
    {
        "categoria": "Transporte",
        "atracao": "Bilhete aeroporto (CDG/Orly)",
        "url": "https://www.iledefrance-mobilites.fr/en/tickets-fares/detail/paris-region-airport-ticket",
        "como_agendar": "Máquinas na estação ou app Bonjour RATP / IDF Mobilités",
        "antecedencia": "No dia; Navigo semanal já inclui se cobrir a semana",
    },
    {
        "categoria": "Museus",
        "atracao": "Louvre — ingressos oficiais",
        "url": "https://ticket.louvre.fr/en",
        "como_agendar": "Site oficial → escolha data, horário e tipo de ingresso (pago ou gratuito)",
        "antecedencia": "Semanas antes; jul–ago 2026 reserva obrigatória; sextas grátis: poucos dias antes",
    },
    {
        "categoria": "Museus",
        "atracao": "Louvre — horários e preços",
        "url": "https://www.louvre.fr/en/visit/hours-admission",
        "como_agendar": "Consulta de horários, preços e dias de entrada gratuita",
        "antecedencia": "Fechado às terças",
    },
    {
        "categoria": "Museus",
        "atracao": "Musée d'Orsay",
        "url": "https://billetterie.musee-orsay.fr/en-GB",
        "como_agendar": "Bilhetes online com horário marcado",
        "antecedencia": "Grátis 1º domingo out–mar; reserve online mesmo quando grátis",
    },
    {
        "categoria": "Monumentos",
        "atracao": "Torre Eiffel — bilhetes oficiais",
        "url": "https://ticket.toureiffel.paris/en",
        "como_agendar": "Escolha data, horário, andar (2º ou topo) e elevador/escadas",
        "antecedencia": "Até 60 dias antes (elevador) ou 30 dias (escadas); esgota rápido",
    },
    {
        "categoria": "Monumentos",
        "atracao": "Torre Eiffel — preços e horários",
        "url": "https://www.toureiffel.paris/en/rates-opening-times",
        "como_agendar": "Referência de tarifas e calendário de abertura",
        "antecedencia": "Consulta antes de comprar",
    },
    {
        "categoria": "Monumentos",
        "atracao": "Arco do Triunfo",
        "url": "https://www.paris-arc-de-triomphe.fr/en/booking/book-a-ticket",
        "como_agendar": "Bilheteria online Centre des Monuments Nationaux",
        "antecedencia": "Grátis 1º domingo nov–mar; quartas abr–set têm preço reduzido",
    },
    {
        "categoria": "Monumentos",
        "atracao": "Sainte-Chapelle",
        "url": "https://billetterie.monuments-nationaux.fr/en/list/monuments-sainte-chapelle",
        "como_agendar": "Ingresso com horário via Monuments Nationaux",
        "antecedencia": "Combine com Notre-Dame (exterior) no Dia 1",
    },
    {
        "categoria": "Monumentos",
        "atracao": "Palácio de Versailles",
        "url": "https://www.chateauversailles.fr/visit/tickets",
        "como_agendar": "Ingressos palácio e jardins no site oficial",
        "antecedencia": "RER C incluso no Navigo; jardins grátis certos dias",
    },
    {
        "categoria": "Parques",
        "atracao": "Disneyland Paris — ingressos",
        "url": "https://www.disneylandparis.com/en-usd/tickets/",
        "como_agendar": "Escolha data, nº de dias e parques (1 ou 2 com Park Hopper)",
        "antecedencia": "Compre cedo; preço dinâmico (€62–130+); RER A incluso no Navigo",
    },
    {
        "categoria": "Parques",
        "atracao": "Disneyland — horários do parque",
        "url": "https://www.disneylandparis.com/en-usd/calendar/",
        "como_agendar": "Calendário de abertura, horários e temporada",
        "antecedencia": "Confira antes do Dia 4 do roteiro",
    },
    {
        "categoria": "Transporte",
        "atracao": "RER A — Paris ↔ Disney (Marne-la-Vallée)",
        "url": "https://www.ratp.fr/en/getting-around/maps/rer-a",
        "como_agendar": "Incluso no Navigo semanal; saia cedo (~45 min de viagem)",
        "antecedencia": "Estação terminus: Marne-la-Vallée—Chessy",
    },
    {
        "categoria": "Utilidades",
        "atracao": "Notre-Dame de Paris (visita)",
        "url": "https://www.notredamedeparis.fr/en/visit/opening-times-and-access",
        "como_agendar": "Reserva gratuita online para entrar na catedral",
        "antecedencia": "Reaberta; reserve horário no site oficial",
    },
    {
        "categoria": "Utilidades",
        "atracao": "eSIM / chip de internet",
        "url": "https://www.airalo.com/europe-esim",
        "como_agendar": "Compre online antes da viagem; ative ao chegar",
        "antecedencia": "Alternativas: Holafly, Orange Holiday",
    },
]

# Mapa item de custo → (texto do link, url)
CUSTOS_LINKS = {
    "Passagem aérea (ida e volta)": ("Comparar voos", "https://www.google.com/travel/flights"),
    "Passe Navigo semanal (zonas 1–5)": ("Navigo semanal", "https://www.iledefrance-mobilites.fr/en/tickets-fares/detail/navigo-weekly-ticket"),
    "Cartão Navigo Découverte (taxa única)": ("Navigo Découverte", "https://www.iledefrance-mobilites.fr/en/tickets-fares/media/navigo-decouverte"),
    "Transfer aeroporto (alternativa ao Navigo)": ("Bilhete aeroporto", "https://www.iledefrance-mobilites.fr/en/tickets-fares/detail/paris-region-airport-ticket"),
    "Disneyland Paris — 1 dia, 1 parque": ("Comprar ingresso", "https://www.disneylandparis.com/en-usd/tickets/"),
    "Disneyland Paris — 1 dia, 2 parques (Park Hopper)": ("Comprar Park Hopper", "https://www.disneylandparis.com/en-usd/tickets/"),
    "Torre Eiffel — 2º andar (elevador)": ("Reservar", "https://ticket.toureiffel.paris/en"),
    "Torre Eiffel — topo (elevador)": ("Reservar topo", "https://ticket.toureiffel.paris/en"),
    "Torre Eiffel — escadas 2º + elevador topo": ("Reservar combo", "https://ticket.toureiffel.paris/en"),
    "Arco do Triunfo — ingresso": ("Reservar", "https://www.paris-arc-de-triomphe.fr/en/booking/book-a-ticket"),
    "Louvre — ingresso pago": ("Reservar Louvre", "https://ticket.louvre.fr/en"),
    "Musée d'Orsay (opcional)": ("Reservar d'Orsay", "https://billetterie.musee-orsay.fr/en-GB"),
    "Versailles (opcional)": ("Reservar Versailles", "https://www.chateauversailles.fr/visit/tickets"),
    "Chip / eSIM internet": ("eSIM Airalo", "https://www.airalo.com/europe-esim"),
}


def set_hyperlink(cell, url, text, size=10):
    cell.value = text
    cell.hyperlink = url
    cell.font = Font(size=size, color="0563C1", underline="single")
    cell.alignment = Alignment(vertical="center", wrap_text=False)


def style_header_row(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def style_data_area(ws, start_row, end_row, cols):
    for row in range(start_row, end_row + 1):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def set_col_widths(ws, widths):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def set_compact_rows(ws, start_row, end_row, height=14.4):
    for row in range(start_row, end_row + 1):
        ws.row_dimensions[row].height = height


def fit_all_rows_on_screen(ws, last_row, zoom=85, row_height=14.4):
    """Compacta linhas e ajusta zoom para exibir o conteúdo inteiro sem rolar."""
    ws.sheet_format.defaultRowHeight = row_height
    set_compact_rows(ws, 1, last_row, row_height)
    ws.sheet_view.zoomScale = zoom
    ws.sheet_view.zoomScaleNormal = zoom
    ws.freeze_panes = None


def build_custos_sheet(wb):
    ws = wb.active
    ws.title = "Custos da Viagem"

    ws.merge_cells("A1:K1")
    title = ws["A1"]
    title.value = "PLANEJAMENTO DE CUSTOS — VIAGEM A PARIS"
    title.font = Font(bold=True, size=14, color="1F4E79")
    title.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:K2")
    info = ws["A2"]
    info.value = "Preencha quantidade e valor unitário. Altere o câmbio global na célula G3."
    info.font = Font(italic=True, color="666666")
    info.alignment = Alignment(horizontal="center")

    ws["F3"].value = "Câmbio global (€→R$):"
    ws["F3"].font = Font(bold=True)
    ws["F3"].alignment = Alignment(horizontal="right", vertical="center")
    ws["G3"].value = 6.20
    ws["G3"].number_format = "0.00"
    ws["G3"].fill = NOTE_FILL
    ws["G3"].border = BORDER
    ws["G3"].alignment = Alignment(horizontal="center")

    headers = [
        "Categoria",
        "Item / Descrição",
        "Qtd. pessoas",
        "Qtd. / dias",
        "Valor unit. (€)",
        "Total estimado (€)",
        "Câmbio (€→R$)",
        "Total estimado (R$)",
        "Valor real pago (R$)",
        "Observações",
        "Link / Reserva",
    ]
    header_row = 4
    for col, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=header)

    items = [
        ("Transporte", "Passagem aérea (ida e volta)", 1, 1, 0, "Comprar com antecedência; preços variam muito por temporada"),
        ("Transporte", "Passe Navigo semanal (zonas 1–5)", 1, 1, 32.40, "Válido de segunda a domingo; inclui metrô, RER, ônibus e Disney"),
        ("Transporte", "Cartão Navigo Découverte (taxa única)", 1, 1, 5.00, "Obrigatório para carregar o passe; traga foto 3×4"),
        ("Transporte", "Transfer aeroporto (alternativa ao Navigo)", 1, 2, 14.00, "Bilhete aeroporto CDG/Orly, se não usar Navigo na chegada"),
        ("Hospedagem", "Hotel / Airbnb (por noite)", 1, 4, 0, "4 noites (chegada dia 1, saída manhã dia 5)"),
        ("Hospedagem", "Taxa de limpeza / city tax", 1, 1, 0, "Verificar no anúncio do hotel"),
        ("Alimentação", "Café da manhã (por dia)", 1, 5, 8.00, "Padaria ou café; ~€8–15/dia"),
        ("Alimentação", "Almoço (por dia)", 1, 5, 15.00, "Bistro ou menu do dia; ~€12–20"),
        ("Alimentação", "Jantar (por dia)", 1, 5, 20.00, "Restaurante médio; ~€15–35"),
        ("Alimentação", "Lanches / mercado / água", 1, 5, 5.00, "Supermercado Carrefour, Monoprix etc."),
        ("Atrações", "Disneyland Paris — 1 dia, 1 parque", 1, 1, 75.00, "Preço dinâmico: €62–130+ conforme data"),
        ("Atrações", "Disneyland Paris — 1 dia, 2 parques (Park Hopper)", 1, 1, 110.00, "Acesso aos dois parques no mesmo dia"),
        ("Atrações", "Torre Eiffel — 2º andar (elevador)", 1, 1, 23.50, "Adulto; reserva online obrigatória"),
        ("Atrações", "Torre Eiffel — topo (elevador)", 1, 1, 36.70, "Adulto; opção mais completa"),
        ("Atrações", "Torre Eiffel — escadas 2º + elevador topo", 1, 1, 28.00, "Melhor custo-benefício para subir ao topo"),
        ("Atrações", "Arco do Triunfo — ingresso", 1, 1, 16.00, "€16 (out–mar) ou €22 (abr–set); quartas €16 na alta"),
        ("Atrações", "Louvre — ingresso pago", 1, 1, 22.00, "Adulto ~€22; veja aba 'Louvre Grátis' para dias gratuitos"),
        ("Atrações", "Musée d'Orsay (opcional)", 1, 1, 16.00, "Grátis 1º domingo do mês (out–mar)"),
        ("Atrações", "Versailles (opcional)", 1, 1, 21.00, "Palácio; RER C incluído no Navigo"),
        ("Extras", "Seguro viagem", 1, 1, 0, "Recomendado; cotar em reais"),
        ("Extras", "Chip / eSIM internet", 1, 1, 0, "Comparar Airalo, Holafly etc."),
        ("Extras", "Souvenirs / compras", 1, 1, 0, "Orçamento livre"),
        ("Extras", "Gorjetas / imprevistos", 1, 1, 50.00, "Reserva de emergência em euros"),
    ]

    start_row = header_row + 1
    for i, (cat, desc, qtd_pessoas, qtd_dias, unit_eur, obs) in enumerate(items):
        row = start_row + i
        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=desc)
        ws.cell(row=row, column=3, value=qtd_pessoas)
        ws.cell(row=row, column=4, value=qtd_dias)
        ws.cell(row=row, column=5, value=unit_eur)
        ws.cell(row=row, column=6, value=f"=C{row}*D{row}*E{row}")
        ws.cell(row=row, column=7, value=f"=${CAMBIO_CELL}")
        ws.cell(row=row, column=8, value=f"=F{row}*G{row}")
        ws.cell(row=row, column=9, value=0)
        ws.cell(row=row, column=10, value=obs)
        if desc in CUSTOS_LINKS:
            link_text, link_url = CUSTOS_LINKS[desc]
            set_hyperlink(ws.cell(row=row, column=11), link_url, link_text)

    end_row = start_row + len(items) - 1
    total_row = end_row + 2

    ws.cell(row=total_row, column=5, value="TOTAIS")
    ws.cell(row=total_row, column=5).font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=f"=SUM(F{start_row}:F{end_row})")
    ws.cell(row=total_row, column=8, value=f"=SUM(H{start_row}:H{end_row})")
    ws.cell(row=total_row, column=9, value=f"=SUM(I{start_row}:I{end_row})")
    for col in (5, 6, 8, 9):
        ws.cell(row=total_row, column=col).fill = TOTAL_FILL
        ws.cell(row=total_row, column=col).font = Font(bold=True)

    per_person_row = total_row + 1
    ws.cell(row=per_person_row, column=5, value="Por pessoa (EUR)")
    ws.cell(row=per_person_row, column=6, value=f"=F{total_row}/MAX(C{start_row},1)")
    ws.cell(row=per_person_row, column=5).font = Font(bold=True, italic=True)

    style_header_row(ws, header_row, len(headers))
    style_data_area(ws, start_row, end_row, len(headers))
    style_data_area(ws, total_row, total_row, len(headers))

    for row in range(start_row, end_row + 1):
        ws.cell(row=row, column=5).number_format = EUR_FMT
        ws.cell(row=row, column=6).number_format = EUR_FMT
        ws.cell(row=row, column=7).number_format = "0.00"
        ws.cell(row=row, column=8).number_format = BRL_FMT
        ws.cell(row=row, column=9).number_format = BRL_FMT

    ws.cell(row=total_row, column=6).number_format = EUR_FMT
    ws.cell(row=total_row, column=8).number_format = BRL_FMT
    ws.cell(row=total_row, column=9).number_format = BRL_FMT
    ws.cell(row=per_person_row, column=6).number_format = EUR_FMT

    ws.freeze_panes = "A5"
    set_col_widths(ws, [14, 38, 12, 12, 14, 16, 14, 18, 18, 42, 18])

    return ws, start_row, end_row


def build_roteiro_sheet(wb):
    ws = wb.create_sheet("Roteiro 5 Dias", 1)

    ws.merge_cells("A1:G1")
    ws["A1"].value = "ROTEIRO DE 5 DIAS — PARIS"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:G2")
    ws["A2"].value = (
        "Modelo: chegada Segunda (Dia 1) → saída Sexta de manhã (Dia 5). "
        "Ajuste datas na coluna B. Links clicáveis na coluna G."
    )
    ws["A2"].font = Font(italic=True, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center", wrap_text=True)

    headers = [
        "Horário", "Data (editável)", "Atividade", "Bairro / local",
        "Transporte", "Custo previsto (€)", "Link / Reserva",
    ]
    header_row = 4
    for col, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=h)
    style_header_row(ws, header_row, len(headers))

    days = [
        {
            "title": "DIA 1 — Chegada e Paris Clássico (Centro)",
            "subtitle": "Segunda-feira | Aclimatação leve após o voo",
            "fill": DAY_FILLS[0],
            "items": [
                ("Manhã", "__/__/____", "Chegada CDG/Orly → hotel (check-in ou deixar bagagem)", "Aeroporto → hotel", "RER B ou Navigo*", "0 / 14*", "Navigo semanal", "https://www.iledefrance-mobilites.fr/en/tickets-fares/detail/navigo-weekly-ticket"),
                ("11h–13h", "__/__/____", "Passeio pelo bairro do hotel + almoço", "Marais / Saint-Germain", "A pé / metrô", "15", "", ""),
                ("14h–16h", "__/__/____", "Notre-Dame (exterior) + Île de la Cité + Sainte-Chapelle (opcional €16)", "Île de la Cité", "Metrô 4", "0–16", "Notre-Dame + Sainte-Chapelle", "https://www.notredamedeparis.fr/en/visit/opening-times-and-access"),
                ("16h30–18h", "__/__/____", "Seine — passeio a pé até Pont Neuf e Louvre (exterior)", "1º arrondissement", "A pé", "0", "Info Louvre", "https://www.louvre.fr/en/visit/hours-admission"),
                ("19h–21h", "__/__/____", "Jantar + descanso", "Bistro perto do hotel", "Metrô", "20–35", "", ""),
            ],
        },
        {
            "title": "DIA 2 — Louvre, Jardins e Torre Eiffel",
            "subtitle": "Terça-feira | Dia mais icônico de Paris",
            "fill": DAY_FILLS[1],
            "items": [
                ("08h30", "__/__/____", "Café da manhã + comprar pão/croissant na padaria", "Perto do hotel", "A pé", "8", "", ""),
                ("09h30–13h", "__/__/____", "Museu do Louvre (reservar horário!) — foco: Mona Lisa, Vênus de Milo, Denon", "Louvre", "Metrô 1 / 7", "22*", "Reservar Louvre", "https://ticket.louvre.fr/en"),
                ("13h–14h", "__/__/____", "Almoço no Carrousel du Louvre ou Tuileries", "1º arr.", "A pé", "15", "", ""),
                ("14h30–16h", "__/__/____", "Jardim das Tuileries + Place de la Concorde", "Tuileries", "A pé", "0", "", ""),
                ("16h30–18h30", "__/__/____", "Torre Eiffel — subir ao topo (reserva online)", "7º arr.", "Metrô 6 / RER C", "28–37", "Reservar Torre", "https://ticket.toureiffel.paris/en"),
                ("19h–21h", "__/__/____", "Pôr do sol no Trocadéro + jantar", "16º arr.", "Metrô 6", "25", "", ""),
            ],
        },
        {
            "title": "DIA 3 — Arco do Triunfo, Champs-Élysées e Montmartre",
            "subtitle": "Quarta-feira | Se for abr–set, Arco €16 (quarta promocional)",
            "fill": DAY_FILLS[2],
            "items": [
                ("09h", "__/__/____", "Café da manhã", "Hotel", "—", "8", "", ""),
                ("09h30–11h", "__/__/____", "Arco do Triunfo — subir ao terraço (vista dos 12 avenidas)", "8º arr.", "Metrô 1 / 2 / 6", "16–22", "Reservar Arco", "https://www.paris-arc-de-triomphe.fr/en/booking/book-a-ticket"),
                ("11h–13h", "__/__/____", "Champs-Élysées a pé até Place de la Concorde", "8º arr.", "A pé", "0", "", ""),
                ("13h–14h", "__/__/____", "Almoço — menu do dia em bistro", "Madeleine / Opéra", "Metrô", "15", "", ""),
                ("14h30–16h", "__/__/____", "Galerias Lafayette ou Printemps (opcional, vistas do terraço grátis)", "9º arr.", "Metrô 7 / 9", "0", "Galerias Lafayette", "https://www.galerieslafayette.com/en/"),
                ("16h30–19h", "__/__/____", "Montmartre: Sacré-Cœur, Place du Tertre, Moulin Rouge (exterior)", "18º arr.", "Metrô 2 / 12", "0", "", ""),
                ("19h30", "__/__/____", "Jantar em Montmartre ou volta ao hotel", "Montmartre", "Metrô", "25", "", ""),
            ],
        },
        {
            "title": "DIA 4 — Disneyland Paris (dia inteiro)",
            "subtitle": "Quinta-feira | Saia cedo — parque abre ~9h30",
            "fill": DAY_FILLS[3],
            "items": [
                ("07h30", "__/__/____", "Café rápido + lanche para levar", "Hotel", "—", "8", "", ""),
                ("08h00", "__/__/____", "RER A → Marne-la-Vallée (Chessy) — ~45 min", "Disney", "RER A (Navigo)", "0**", "Mapa RER A", "https://www.ratp.fr/en/getting-around/maps/rer-a"),
                ("09h30–13h", "__/__/____", "Disneyland Park — manhã: Big Thunder, Phantom Manor, Pirates", "Parque 1", "A pé", "75–110***", "Comprar ingresso", "https://www.disneylandparis.com/en-usd/tickets/"),
                ("13h–14h", "__/__/____", "Almoço dentro do parque (ou lanche)", "Disney", "—", "15–25", "Calendário parque", "https://www.disneylandparis.com/en-usd/calendar/"),
                ("14h–17h", "__/__/____", "Walt Disney Studios (se Park Hopper) ou continuar Parque 1", "Parque 2 / 1", "A pé", "incl.***", "", ""),
                ("17h–20h", "__/__/____", "Desfile / show + jantar no parque", "Disney", "—", "20–30", "", ""),
                ("21h", "__/__/____", "Retorno a Paris (RER A)", "Hotel", "RER A", "0", "", ""),
            ],
        },
        {
            "title": "DIA 5 — Últimas compras e partida",
            "subtitle": "Sexta-feira | Check-out de manhã, voo à tarde/noite",
            "fill": DAY_FILLS[4],
            "items": [
                ("08h", "__/__/____", "Check-out + guardar bagagem no hotel (se permitido)", "Hotel", "—", "0", "", ""),
                ("08h30–10h30", "__/__/____", "Marché ou padaria + últimas compras (Le Marais ou Saint-Germain)", "4º / 6º arr.", "Metrô", "10–30", "", ""),
                ("10h30–12h", "__/__/____", "Musée d'Orsay (opcional) OU passeio livre pelo Sena", "7º arr.", "RER C / metrô", "0–16", "Reservar d'Orsay", "https://billetterie.musee-orsay.fr/en-GB"),
                ("12h–13h30", "__/__/____", "Almoço de despedida", "Perto do hotel", "Metrô", "15–25", "", ""),
                ("14h", "__/__/____", "Retorno ao hotel → pegar bagagem", "Hotel", "Metrô", "0", "", ""),
                ("15h–16h", "__/__/____", "Transfer para aeroporto (CDG ~1h / Orly ~45 min)", "Aeroporto", "RER B / Navigo", "0 / 14*", "Bilhete aeroporto", "https://www.iledefrance-mobilites.fr/en/tickets-fares/detail/paris-region-airport-ticket"),
            ],
        },
    ]

    current_row = header_row + 1
    for day in days:
        ws.merge_cells(f"A{current_row}:G{current_row}")
        title_cell = ws.cell(row=current_row, column=1, value=day["title"])
        title_cell.font = Font(bold=True, size=12, color="1F4E79")
        title_cell.fill = day["fill"]
        title_cell.alignment = Alignment(vertical="center")
        for col in range(1, 8):
            ws.cell(row=current_row, column=col).border = BORDER
            if col > 1:
                ws.cell(row=current_row, column=col).fill = day["fill"]
        current_row += 1

        ws.merge_cells(f"A{current_row}:G{current_row}")
        sub_cell = ws.cell(row=current_row, column=1, value=day["subtitle"])
        sub_cell.font = Font(italic=True, color="444444")
        sub_cell.fill = day["fill"]
        for col in range(1, 8):
            ws.cell(row=current_row, column=col).border = BORDER
            if col > 1:
                ws.cell(row=current_row, column=col).fill = day["fill"]
        current_row += 1

        for item in day["items"]:
            link_text = item[6] if len(item) > 6 else ""
            link_url = item[7] if len(item) > 7 else ""
            for col, val in enumerate(item[:6], start=1):
                cell = ws.cell(row=current_row, column=col, value=val)
                cell.border = BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            link_cell = ws.cell(row=current_row, column=7)
            link_cell.border = BORDER
            if link_text and link_url:
                set_hyperlink(link_cell, link_url, link_text)
            current_row += 1

        current_row += 1

    notes_start = current_row
    footnotes = [
        "NOTAS DO ROTEIRO",
        "",
        "* Navigo semanal (€32,40): se a viagem cair segunda–domingo, cobre aeroporto, Disney e todo o transporte.",
        "** Disney incluído no Navigo (RER A até Marne-la-Vallée). Shuttle VEA Disney NÃO está incluso.",
        "*** Ingresso Disney: reserve com antecedência; 1 parque ~€75 ou 2 parques (Hopper) ~€110 (preços dinâmicos).",
        "* Louvre €22 — GRÁTIS na 1ª sexta do mês (18h–21h45), exceto jul/ago. Veja aba 'Louvre Grátis'.",
        "Arco do Triunfo: gratuito no 1º domingo nov–mar. Se coincidir, encaixe no Dia 3 ou 5.",
        "Todos os links oficiais estão na aba 'Links e Reservas' e na coluna G acima.",
        "Dica jet lag: Dia 1 propositalmente leve. Não empilhe Disney + Louvre no mesmo dia.",
    ]
    for i, line in enumerate(footnotes):
        row = notes_start + i
        ws.merge_cells(f"A{row}:G{row}")
        cell = ws.cell(row=row, column=1, value=line)
        if line == "NOTAS DO ROTEIRO":
            cell.font = Font(bold=True, size=11, color="1F4E79")
        elif line.startswith("*"):
            cell.fill = NOTE_FILL
        cell.alignment = Alignment(wrap_text=True)

    ws.freeze_panes = "A5"
    set_col_widths(ws, [12, 14, 42, 22, 18, 16, 20])

    return ws


def build_louvre_sheet(wb):
    ws = wb.create_sheet("Louvre Grátis")

    ws.merge_cells("A1:E1")
    ws["A1"].value = "MUSEU DO LOUVRE — ENTRADA GRATUITA"
    ws["A1"].font = Font(bold=True, size=12, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 18

    link_row = 2
    ws.cell(row=link_row, column=1, value="Links:")
    ws.cell(row=link_row, column=1).font = Font(bold=True, size=9)
    set_hyperlink(ws.cell(row=link_row, column=2), "https://ticket.louvre.fr/en", "Reservar ingresso")
    set_hyperlink(ws.cell(row=link_row, column=3), "https://www.louvre.fr/en/visit/hours-admission", "Horários e preços")
    set_hyperlink(
        ws.cell(row=link_row, column=4),
        "https://www.paris-arc-de-triomphe.fr/en/booking/book-a-ticket",
        "Arco do Triunfo",
    )
    set_hyperlink(
        ws.cell(row=link_row, column=5),
        "https://billetterie.musee-orsay.fr/en-GB",
        "Musée d'Orsay",
    )
    ws.row_dimensions[link_row].height = 16

    notes_row = 3
    left_notes = (
        "IMPORTANTE: Não é grátis no 1º domingo (política encerrada).\n"
        "Reserva online OBRIGATÓRIA, mesmo nos dias grátis.\n\n"
        "Grátis para TODOS:\n"
        "• 1ª sexta do mês, 18h–21h45 (exc. jul/ago)\n"
        "• 14/07 — Bastilha (se aberto)"
    )
    right_notes = (
        "Grátis o ano todo (com documento):\n"
        "• Menores de 18 anos\n"
        "• EEE menores de 26 anos\n"
        "• PCD + acompanhante\n\n"
        "Ingresso pago: ~€22 adulto"
    )
    note_font = Font(size=9)

    ws.merge_cells(f"A{notes_row}:B{notes_row}")
    left_cell = ws.cell(row=notes_row, column=1, value=left_notes)
    left_cell.font = note_font
    left_cell.alignment = Alignment(wrap_text=True, vertical="top")
    left_cell.fill = NOTE_FILL
    left_cell.border = BORDER

    ws.merge_cells(f"C{notes_row}:E{notes_row}")
    right_cell = ws.cell(row=notes_row, column=3, value=right_notes)
    right_cell.font = note_font
    right_cell.alignment = Alignment(wrap_text=True, vertical="top")
    right_cell.fill = NOTE_FILL
    right_cell.border = BORDER
    ws.row_dimensions[notes_row].height = 68

    header_row = 4
    headers = ["Data", "Dia", "Horário", "Observação", "Reservar"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    free_dates = [
        ("02/01/2026", "Sexta", "18h–21h45", "1ª sexta do mês"),
        ("06/02/2026", "Sexta", "18h–21h45", "1ª sexta do mês"),
        ("06/03/2026", "Sexta", "18h–21h45", "1ª sexta do mês"),
        ("03/04/2026", "Sexta", "18h–21h45", "1ª sexta do mês"),
        ("—", "—", "—", "Maio: fechado 01/05 (Dia do Trabalho)"),
        ("05/06/2026", "Sexta", "18h–21h45", "Última sexta grátis antes do verão"),
        ("—", "—", "—", "Julho e agosto: sem entrada gratuita"),
        ("04/09/2026", "Sexta", "18h–21h45", "Retorno após pausa de verão"),
        ("02/10/2026", "Sexta", "18h–21h45", "1ª sexta do mês"),
        ("06/11/2026", "Sexta", "18h–21h45", "1ª sexta do mês"),
        ("04/12/2026", "Sexta", "18h–21h45", "Última do ano"),
        ("14/07/2026", "Terça", "Horário normal", "Bastilha — museu FECHADO às terças"),
    ]

    data_start = header_row + 1
    for i, row_data in enumerate(free_dates):
        row = data_start + i
        for col, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = Alignment(vertical="center", wrap_text=False)
        if row_data[0] != "—":
            set_hyperlink(
                ws.cell(row=row, column=5),
                "https://ticket.louvre.fr/en",
                "Reservar",
                size=9,
            )

    style_header_row(ws, header_row, 5)
    style_data_area(ws, data_start, data_start + len(free_dates) - 1, 5)

    tip_row = data_start + len(free_dates)
    ws.merge_cells(f"A{tip_row}:E{tip_row}")
    ws[f"A{tip_row}"].value = (
        "Dica: Arco do Triunfo e Musée d'Orsay grátis no 1º domingo (nov–mar / out–mar). "
        "Use os links da linha 2."
    )
    ws[f"A{tip_row}"].fill = NOTE_FILL
    ws[f"A{tip_row}"].font = Font(size=9)
    ws[f"A{tip_row}"].alignment = Alignment(vertical="center", wrap_text=False)
    ws.row_dimensions[tip_row].height = 16

    set_col_widths(ws, [12, 8, 12, 36, 14])
    fit_all_rows_on_screen(ws, tip_row, zoom=90, row_height=14.4)
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[link_row].height = 16
    ws.row_dimensions[notes_row].height = 68
    ws.row_dimensions[tip_row].height = 16


def build_links_sheet(wb):
    ws = wb.create_sheet("Links e Reservas", 2)

    ws.merge_cells("A1:E1")
    ws["A1"].value = "LINKS OFICIAIS — RESERVAS E INGRESSOS"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:E2")
    ws["A2"].value = "Clique nos links azul sublinhado. Use sempre os sites oficiais para evitar fraudes."
    ws["A2"].font = Font(italic=True, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center", wrap_text=True)

    headers = ["Categoria", "Atração / serviço", "Link (clique)", "Como agendar", "Antecedência recomendada"]
    header_row = 4
    for col, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=h)
    style_header_row(ws, header_row, len(headers))

    data_start = header_row + 1
    for i, entry in enumerate(BOOKING_LINKS):
        row = data_start + i
        ws.cell(row=row, column=1, value=entry["categoria"])
        ws.cell(row=row, column=2, value=entry["atracao"])
        set_hyperlink(ws.cell(row=row, column=3), entry["url"], "Abrir site oficial")
        ws.cell(row=row, column=4, value=entry["como_agendar"])
        ws.cell(row=row, column=5, value=entry["antecedencia"])

    end_row = data_start + len(BOOKING_LINKS) - 1
    style_data_area(ws, data_start, end_row, len(headers))

    note_row = end_row + 2
    ws.merge_cells(f"A{note_row}:E{note_row}")
    ws[f"A{note_row}"].value = (
        "Atenção: evite sites espelho ou vendedores na rua. "
        "Louvre, Torre Eiffel e Disney só vendem ingressos válidos pelos sites oficiais listados acima."
    )
    ws[f"A{note_row}"].fill = NOTE_FILL
    ws[f"A{note_row}"].alignment = Alignment(wrap_text=True)

    ws.freeze_panes = "A5"
    set_col_widths(ws, [14, 32, 18, 42, 36])


def build_resumo_sheet(wb, start_row, end_row):
    ws = wb.create_sheet("Resumo e Dicas")

    ws.merge_cells("A1:C1")
    ws["A1"].value = "REFERÊNCIA RÁPIDA DE PREÇOS (2026)"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")

    ref_data = [
        ("Item", "Preço adulto (€)", "Link oficial"),
        ("Navigo semanal (zonas 1–5)", "32,40", "iledefrance-mobilites.fr"),
        ("Cartão Navigo Découverte", "5,00", "iledefrance-mobilites.fr"),
        ("Disneyland — 1 dia, 1 parque", "62 – 130+", "disneylandparis.com"),
        ("Disneyland — 1 dia, 2 parques", "89 – 140+", "disneylandparis.com"),
        ("Torre Eiffel — 2º andar elevador", "23,50", "ticket.toureiffel.paris"),
        ("Torre Eiffel — topo elevador", "36,70", "ticket.toureiffel.paris"),
        ("Torre Eiffel — escadas + topo", "28,00", "ticket.toureiffel.paris"),
        ("Arco do Triunfo", "16 / 22", "paris-arc-de-triomphe.fr"),
        ("Louvre (pago)", "22,00", "ticket.louvre.fr"),
        ("Alimentação/dia (econômico)", "30 – 45", "—"),
        ("Alimentação/dia (confortável)", "50 – 80", "—"),
    ]

    ref_links = {
        4: "https://www.iledefrance-mobilites.fr/en/tickets-fares/detail/navigo-weekly-ticket",
        5: "https://www.iledefrance-mobilites.fr/en/tickets-fares/media/navigo-decouverte",
        6: "https://www.disneylandparis.com/en-usd/tickets/",
        7: "https://www.disneylandparis.com/en-usd/tickets/",
        8: "https://ticket.toureiffel.paris/en",
        9: "https://ticket.toureiffel.paris/en",
        10: "https://ticket.toureiffel.paris/en",
        11: "https://www.paris-arc-de-triomphe.fr/en/booking/book-a-ticket",
        12: "https://ticket.louvre.fr/en",
    }

    for i, row_data in enumerate(ref_data, start=3):
        for col, val in enumerate(row_data, start=1):
            ws.cell(row=i, column=col, value=val)
        if i in ref_links:
            set_hyperlink(ws.cell(row=i, column=3), ref_links[i], row_data[2])

    style_header_row(ws, 3, 3)
    style_data_area(ws, 4, 3 + len(ref_data) - 1, 3)

    tips_start = 3 + len(ref_data) + 2
    ws.merge_cells(f"A{tips_start}:C{tips_start}")
    ws[f"A{tips_start}"].value = "DICAS DE ECONOMIA"
    ws[f"A{tips_start}"].font = Font(bold=True, size=12, color="1F4E79")

    tips = [
        "Compre passagens aéreas com 2–3 meses de antecedência.",
        "Alinhe as datas com o Navigo semanal (segunda–domingo) para maximizar o passe.",
        "Disneyland: evite fins de semana e férias escolares europeias para preços menores.",
        "Louvre grátis: reserve assim que abrirem os horários (poucos dias antes da data).",
        "Prefira menu do dia (formule) no almoço — geralmente €14–18 com prato + sobremesa.",
        "Mercados e padarias reduzem muito o custo de alimentação.",
        "Arco do Triunfo: visite numa quarta entre abril e setembro e pague €16 em vez de €22.",
        "Todos os links de reserva estão na aba 'Links e Reservas'.",
    ]
    for i, tip in enumerate(tips, start=tips_start + 1):
        ws.merge_cells(f"A{i}:C{i}")
        ws[f"A{i}"].value = f"• {tip}"

    summary_start = tips_start + len(tips) + 2
    summary_rows = [
        ("Câmbio usado na planilha:", "=CambioEUR", "Altere na aba Custos (célula G3)"),
        ("Total estimado EUR:", f"='Custos da Viagem'!F{end_row + 2}", ""),
        ("Total estimado BRL:", f"='Custos da Viagem'!H{end_row + 2}", ""),
    ]
    for i, (label, formula, note) in enumerate(summary_rows):
        row = summary_start + i
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=formula)
        ws.cell(row=row, column=3, value=note)
        for col in range(1, 4):
            cell = ws.cell(row=row, column=col)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.cell(row=row, column=1).font = Font(bold=True)

    ws.cell(row=summary_start, column=2).number_format = "0.00"
    ws.cell(row=summary_start + 1, column=2).number_format = EUR_FMT
    ws.cell(row=summary_start + 2, column=2).number_format = BRL_FMT

    set_col_widths(ws, [32, 18, 50])


def main():
    wb = Workbook()
    _, start_row, end_row = build_custos_sheet(wb)
    wb.defined_names.add(
        DefinedName("CambioEUR", attr_text=f"'Custos da Viagem'!${CAMBIO_CELL}")
    )
    build_roteiro_sheet(wb)
    build_links_sheet(wb)
    build_louvre_sheet(wb)
    build_resumo_sheet(wb, start_row, end_row)
    wb.save(OUTPUT)
    print(f"Planilha criada: {OUTPUT}")


if __name__ == "__main__":
    main()
